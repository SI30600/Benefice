"""OneDrive Excel integration via Microsoft Graph API.

Uses Microsoft OAuth 2.0 (Authorization Code flow) and stores refresh tokens
in MongoDB. The Excel file lives in the user's personal OneDrive at
ONEDRIVE_FILE_PATH and gets one row appended per devis.
"""
from __future__ import annotations

import io
import os
import secrets
from datetime import datetime, timezone
from typing import Optional

import msal
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
SCOPES = ["Files.ReadWrite", "User.Read"]

# Excel column headers (1 row = 1 devis)
HEADERS = [
    "Date", "Nom client", "Adresse", "Code postal", "Ville",
    "Téléphone", "Email", "Type machine",
    "Composants (détail)", "Coût pièces (€)", "Pièces facturées (€)",
    "Livraison achat LBC (€)",
    "Licence Windows (€)", "Prestation (libellé)", "Prestation (€)",
    "Déplacement (zone)", "Déplacement (€)",
    "Total facturé (€)", "Taxe LBC (€)", "URSSAF Articles 13% (€)",
    "URSSAF Prestations 23% (€)", "Bénéfice net (€)", "Marge (%)",
]


def _msal_app() -> msal.ConfidentialClientApplication:
    return msal.ConfidentialClientApplication(
        client_id=os.environ["MS_CLIENT_ID"],
        client_credential=os.environ["MS_CLIENT_SECRET"],
        authority=os.environ["MS_AUTHORITY"],
    )


def build_auth_url(state: str) -> str:
    return _msal_app().get_authorization_request_url(
        scopes=SCOPES,
        redirect_uri=os.environ["MS_REDIRECT_URI"],
        state=state,
    )


def exchange_code_for_token(code: str) -> dict:
    result = _msal_app().acquire_token_by_authorization_code(
        code=code,
        scopes=SCOPES,
        redirect_uri=os.environ["MS_REDIRECT_URI"],
    )
    if "access_token" not in result:
        raise RuntimeError(f"Microsoft auth failed: {result.get('error_description', result)}")
    return result


def refresh_access_token(refresh_token: str) -> dict:
    result = _msal_app().acquire_token_by_refresh_token(
        refresh_token=refresh_token,
        scopes=SCOPES,
    )
    if "access_token" not in result:
        raise RuntimeError(f"Refresh failed: {result.get('error_description', result)}")
    return result


def get_user_profile(access_token: str) -> dict:
    r = requests.get(
        f"{GRAPH_BASE}/me",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


# ---- Excel file handling ----------------------------------------------------

def _new_workbook_bytes() -> bytes:
    """Create a brand new workbook with headers + a named table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Devis"
    ws.append(HEADERS)

    # Header style
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid")
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set sensible column widths
    widths = [12, 22, 28, 11, 16, 16, 24, 12, 40, 14, 16, 16, 18, 18, 18, 14, 16, 14, 18, 20, 16, 10]
    for i, w in enumerate(widths[: len(HEADERS)], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Add a real Excel table so future appends are easy
    last_col_letter = ws.cell(row=1, column=len(HEADERS)).column_letter
    table = Table(displayName="DevisTable", ref=f"A1:{last_col_letter}2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    # An Excel Table needs at least one data row to be valid → empty placeholder
    ws.append([""] * len(HEADERS))
    ws.add_table(table)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ensure_file_on_onedrive(access_token: str, path: str) -> None:
    """If the file doesn't exist, upload an empty templated workbook."""
    headers = {"Authorization": f"Bearer {access_token}"}
    r = requests.get(f"{GRAPH_BASE}/me/drive/root:/{path}", headers=headers, timeout=15)
    if r.status_code == 200:
        return
    if r.status_code != 404:
        r.raise_for_status()

    # Create
    content = _new_workbook_bytes()
    upload = requests.put(
        f"{GRAPH_BASE}/me/drive/root:/{path}:/content",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        data=content,
        timeout=60,
    )
    upload.raise_for_status()


def append_devis_row(access_token: str, path: str, row: list) -> dict:
    """Download → append → upload approach.

    Simpler than Graph Workbook tables (no session juggling, no race issues
    with the placeholder row added at file creation). For a single user with a
    small file, the round-trip cost is negligible.
    """
    _ensure_file_on_onedrive(access_token, path)

    headers = {"Authorization": f"Bearer {access_token}"}
    download = requests.get(
        f"{GRAPH_BASE}/me/drive/root:/{path}:/content",
        headers=headers,
        timeout=30,
    )
    download.raise_for_status()

    wb = load_workbook(io.BytesIO(download.content))
    ws = wb.active

    # Drop the placeholder empty row if present (row 2 with all blanks)
    if ws.max_row >= 2:
        second = [ws.cell(row=2, column=c).value for c in range(1, len(HEADERS) + 1)]
        if all((v is None or v == "") for v in second):
            ws.delete_rows(2, 1)

    ws.append(row)

    # Re-extend the table range to include the new row
    for tbl in ws.tables.values():
        first_cell, _ = tbl.ref.split(":")
        last_col_letter = ws.cell(row=1, column=len(HEADERS)).column_letter
        tbl.ref = f"{first_cell}:{last_col_letter}{ws.max_row}"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    upload = requests.put(
        f"{GRAPH_BASE}/me/drive/root:/{path}:/content",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        data=out.getvalue(),
        timeout=60,
    )
    upload.raise_for_status()
    return upload.json()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def gen_state() -> str:
    return secrets.token_urlsafe(24)
